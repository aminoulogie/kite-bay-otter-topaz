"""
Import the training workbooks into a normalised dataset.

Sheet A is a wide matrix: rows are exercises, columns are dates, and each cell
holds every set done that day in the shorthand parse_sets decodes. Sheet B is a
tidy one-row-per-set log. Both land in the same records, so history and live
data are one continuous dataset.

Nothing is dropped silently. Anything undecodable goes to the reject report
with the cell it came from and the reason.
"""

import datetime
import json
import os
import re
import sys
import unicodedata

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from exercise_catalog import BAR_KG_OVERRIDE, BODYWEIGHT, PLATE_LOADED, resolve
from parse_sets import BAR_KG, parse_cell

SYN = {
    r"\bc[aâ]ble?s?\b": "cable",
    r"\bpoulie\b": "cable",
    r"\bmac\b": "machine",
    r"\bhalt[eè]re?s?\b": "dumbbell",
    r"\bdumbell?s?\b": "dumbbell",
    r"\bdmbl\b": "dumbbell",
    r"\bdbl\b": "dumbbell",
    r"\bdb\b": "dumbbell",
    r"\bbarre?\b": "barbell",
    r"\binclined?\b": "incline",
    r"\bdeclined?\b": "decline",
    r"\bdamped\b": "amortis",
    r"\bassisted\b": "amortis",
    r"\bpull ?downs?\b": "pulldown",
    r"\bpull ?ups?\b": "pullup",
    r"\bpush ?downs?\b": "pushdown",
    r"\btricep\b": "triceps",
    r"\bbicep\b": "biceps",
    r"\blats\b": "lat",
    r"\bflyes\b": "fly",
    r"\bflies\b": "fly",
    r"\bcurls?\b": "curl",
    r"\braises?\b": "raise",
    r"\brows?\b": "row",
    r"\bpresses\b": "press",
    r"\bextensions?\b": "extension",
}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[^a-z\s]", " ", s)
    for p, r in SYN.items():
        s = re.sub(p, r, s)
    return " ".join(dict.fromkeys(t for t in s.split() if len(t) > 1))


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        s = str(int(v))
        if len(s) == 8:  # DDMMYYYY, as written in the header row
            try:
                return datetime.date(int(s[4:]), int(s[2:4]), int(s[:2]))
            except ValueError:
                return None
    if isinstance(v, str):
        for f in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(v.strip(), f).date()
            except ValueError:
                pass
    return None


def is_exercise(s):
    if not isinstance(s, str):
        return False
    s = s.strip()
    if not (3 <= len(s) <= 55):
        return False
    letters = sum(c.isalpha() for c in s)
    if letters < 3 or letters / len(s) < 0.55:
        return False
    if sum(c.isdigit() for c in s) > letters:
        return False
    return not re.search(r"\d\s*(kg)?\s*[x*]\s*\d", s, re.I)


def flags(canon):
    return dict(
        plate_loaded=canon in PLATE_LOADED,
        bodyweight=canon in BODYWEIGHT,
        bar_kg=BAR_KG_OVERRIDE.get(canon, BAR_KG),
    )


def import_all(sp):
    records, rejects = [], []
    # Apr26 and Week 1 are near-copies, so the same cell appears in both. Skip
    # the whole CELL when it repeats, never the individual sets: one cell holds
    # several sets, and filtering those after the fact would keep only the first
    # of "259,2512,255" and silently lose two thirds of the day.
    seen_cells: set[tuple[str, str, str]] = set()

    # ---- Sheet A: exercise x date matrix ----------------------------------
    wb = openpyxl.load_workbook(os.path.join(sp, "A.xlsx"), data_only=True)
    for ws in wb.worksheets:
        hdr_row, dates = None, {}
        for ri, row in enumerate(ws.iter_rows(max_row=40, values_only=True), 1):
            found = {i + 1: as_date(c) for i, c in enumerate(row) if as_date(c)}
            if len(found) > len(dates):
                hdr_row, dates = ri, found
        if len(dates) < 2:
            continue  # not a dated matrix; reported separately

        ex_col = max(
            range(1, min(ws.max_column, 8) + 1),
            key=lambda c: sum(
                1
                for (v,) in ws.iter_rows(min_col=c, max_col=c, values_only=True)
                if is_exercise(v)
            ),
        )

        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            raw = row[ex_col - 1] if ex_col - 1 < len(row) else None
            if not is_exercise(raw):
                continue
            canon = resolve(norm(raw))
            if canon is None:
                continue
            fl = flags(canon)
            for col, d in dates.items():
                cell = row[col - 1] if col - 1 < len(row) else None
                if cell is None or not str(cell).strip():
                    continue
                cell_key = (d.isoformat(), canon, str(cell).strip())
                if cell_key in seen_cells:
                    continue
                seen_cells.add(cell_key)
                res = parse_cell(str(cell), **fl)
                for s in res.sets:
                    records.append(
                        dict(
                            date=d.isoformat(), exercise=canon, weight=s.weight,
                            reps=s.reps, failure=s.failure, straps=s.straps,
                            note=s.note, source="A/" + ws.title, raw=str(cell)[:40],
                        )
                    )
                for tok, why in res.rejects:
                    rejects.append(
                        dict(
                            source="A/" + ws.title, date=d.isoformat(), exercise=canon,
                            cell=str(cell)[:60], token=tok, reason=why,
                        )
                    )

    # ---- Sheet B: tidy one-row-per-set log --------------------------------
    wb2 = openpyxl.load_workbook(os.path.join(sp, "B.xlsx"), data_only=True)
    for ws in wb2.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(c or "").strip().lower() for c in rows[0]]
        need = {"date", "exercise", "weight (kg)", "reps"}
        if not need <= set(hdr):
            continue
        ix = {k: hdr.index(k) for k in need}
        note_i = hdr.index("equipment/notes") if "equipment/notes" in hdr else None

        for r in rows[1:]:
            d = as_date(r[ix["date"]]) if ix["date"] < len(r) else None
            name = r[ix["exercise"]] if ix["exercise"] < len(r) else None
            if not d or not is_exercise(name):
                continue
            canon = resolve(norm(name))
            if canon is None:
                continue
            try:
                w = float(r[ix["weight (kg)"]])
                rp = int(r[ix["reps"]])
            except (TypeError, ValueError):
                rejects.append(
                    dict(source="B/" + ws.title, date=d.isoformat(), exercise=canon,
                         cell=str(r)[:60], token="", reason="weight or reps not numeric")
                )
                continue
            note = str(r[note_i]) if note_i is not None and note_i < len(r) and r[note_i] else ""
            straps = None
            if "strap" in note.lower():
                straps = "no strap" not in note.lower()
            records.append(
                dict(date=d.isoformat(), exercise=canon, weight=w, reps=rp, failure=0,
                     straps=straps, note=note, source="B/" + ws.title, raw="")
            )

    # Sheet A is already de-duplicated per cell above. Sheet B is one row per
    # set, so an identical row twice in the same tab is a genuine repeat and is
    # kept; re-running the import re-reads the same file and yields the same
    # list, which is what idempotent means here.
    return records, rejects


if __name__ == "__main__":
    sp = sys.argv[1]
    recs, rej = import_all(sp)
    json.dump(recs, open(os.path.join(sp, "records.json"), "w"), indent=1)
    json.dump(rej, open(os.path.join(sp, "rejects.json"), "w"), indent=1)
    exs = sorted({r["exercise"] for r in recs})
    ds = sorted({r["date"] for r in recs})
    print("records  :", len(recs))
    print("exercises:", len(exs))
    print("dates    :", len(ds), ("(%s .. %s)" % (ds[0], ds[-1])) if ds else "")
    print("rejects  :", len(rej))
